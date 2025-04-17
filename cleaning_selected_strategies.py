import os
import shutil
import pandas as pd

def copy_selected_files():
    # Define paths
    selected_strategies_dir = r"C:\Users\33698\Desktop\VIX_Trading_Strategy\Selected_Strategies"
    strategies_csv_path = os.path.join(selected_strategies_dir, "Selected_Strategies_Summary.csv")
    best_new_strategies_dir = r"C:\Users\33698\Desktop\VIX_Trading_Strategy\VIX_put_spread_strategies_maturity_adjusted\Best_New_Strategies"
    
    # Read CSV file
    try:
        df = pd.read_csv(strategies_csv_path, usecols=[0], header=None, names=['Filename'])
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        return
    
    # Loop through filenames in column A
    for filename in df['Filename']:
        summary_filename = f"Summary_{filename}"
        source_path = os.path.join(best_new_strategies_dir, summary_filename)
        destination_path = os.path.join(selected_strategies_dir, summary_filename)
        
        # Check if file exists in Best_New_Strategies directory
        if os.path.exists(source_path):
            try:
                shutil.copy2(source_path, destination_path)
                print(f"Copied: {summary_filename}")
            except Exception as e:
                print(f"Error copying {summary_filename}: {e}")
        else:
            print(f"File not found: {summary_filename}")

if __name__ == "__main__":
    copy_selected_files()

import openpyxl

def split_csv_by_dte():
    # Define the directory
    directory = r"C:\Users\33698\Desktop\VIX_Trading_Strategy\Selected_Strategies"
    
    # Get all CSV files except 'Selected_Strategies.csv'
    csv_files = [f for f in os.listdir(directory) if f.endswith('.csv') and f != "Selected_Strategies_Summary.csv"]
    
    for file in csv_files:
        file_path = os.path.join(directory, file)
        df = pd.read_csv(file_path)
        
        if 'DTE' not in df.columns:
            print(f"Skipping {file}: No 'DTE' column found.")
            continue
        
        try:
            df['DTE'] = df['DTE'].astype(str)
            dte_values = set()
            
            for dte_list in df['DTE'].dropna():
                dte_values.update([int(dte.strip()) for dte in dte_list.split(';') if dte.strip().isdigit() and int(dte.strip()) > 4])
            
            dte_values = sorted(dte_values)  # Sort chronologically
            
            # Create a new Excel file
            excel_path = os.path.join(directory, file.replace('.csv', '.xlsx').replace('Summary_', 'Selected_'))
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Write the original sheet
                df.to_excel(writer, sheet_name='Original', index=False)
                
                for dte in dte_values:
                    filtered_df = df[df['DTE'].apply(lambda x: any(str(dte) == item.strip() for item in x.split(';')))]
                    if not filtered_df.empty:
                        filtered_df.to_excel(writer, sheet_name=f'DTE_{dte}', index=False)
                
            print(f"Processed {file}: Created {len(dte_values)} sheets in {excel_path}")
        except Exception as e:
            print(f"Error processing {file}: {e}")
    
    # Delete files that do not start with 'Selected_' AFTER generating the new files
    for file in os.listdir(directory):
        file_path = os.path.join(directory, file)
        if os.path.isfile(file_path) and not file.startswith("Selected_"):
            os.remove(file_path)
            print(f"Deleted: {file_path}")

if __name__ == "__main__":
    split_csv_by_dte()


    import os
import pandas as pd
import openpyxl

def process_selected_files():
    selected_dir = r"C:\Users\33698\Desktop\VIX_Trading_Strategy\Selected_Strategies"
    strategies_dir = r"C:\Users\33698\Desktop\VIX_Trading_Strategy\VIX_put_spread_strategies_maturity_adjusted"
    
    # Get all Selected_ files except 'Selected_Strategies_Summary.csv'
    selected_files = [f for f in os.listdir(selected_dir) if f.startswith("Selected_") and f.endswith(".xlsx") and f != "Selected_Strategies_Summary.csv"]
    
    for file in selected_files:
        file_path = os.path.join(selected_dir, file)
        
        # Extract strike and threshold from filename
        parts = file.replace("Selected_", "").replace(".xlsx", "").split("_")
        try:
            threshold_index = parts.index("threshold") + 1
            strike_index = parts.index("strike") + 1
            threshold = parts[threshold_index]
            strike = parts[strike_index]
        except (ValueError, IndexError):
            print(f"Skipping {file}: Unable to extract strike and threshold.")
            continue
        
        # Define corresponding strategy file path
        strike_folder = f"Strike_{strike}"
        strategy_file = f"vix_put_spread_results_threshold_{threshold}_strike_{strike}.csv"
        strategy_path = os.path.join(strategies_dir, strike_folder, strategy_file)
        
        if not os.path.exists(strategy_path):
            print(f"Skipping {file}: Corresponding strategy file not found: {strategy_path}")
            continue
        
        # Load strategy data
        strategy_df = pd.read_csv(strategy_path)
        required_columns = {"DTE", "RETURN", "SPREAD_COST", "EXPIRY_VALUE"}
        if not required_columns.issubset(strategy_df.columns):
            print(f"Skipping {file}: Required columns missing in strategy file.")
            continue
        
        # Open selected file and update each sheet
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            if sheet_name == "Original":
                continue
            
            try:
                ws = wb[sheet_name]
                dte_value = int(sheet_name.replace("DTE_", ""))
                
                # Clear existing data
                ws.delete_rows(1, ws.max_row)
                
                # Filter strategy data for this DTE
                dte_filtered = strategy_df[strategy_df["DTE"] == dte_value]
                if dte_filtered.empty:
                    print(f"No data for DTE {dte_value} in {strategy_file}")
                    continue
                
                # Calculate required metrics
                num_trades = len(dte_filtered)
                total_return = dte_filtered["RETURN"].sum()
                max_drawdown = dte_filtered[dte_filtered["RETURN"] < 0]["RETURN"].sum()
                winrate = (dte_filtered["RETURN"] > 0).mean() * 100
                
                risk_reward_ratio = abs(total_return / max_drawdown) if max_drawdown != 0 else None
                
                spread_cost_q1 = dte_filtered["SPREAD_COST"].quantile(0.25)
                spread_cost_avg = dte_filtered["SPREAD_COST"].mean()
                spread_cost_q3 = dte_filtered["SPREAD_COST"].quantile(0.75)
                
                expiry_value_q1 = dte_filtered["EXPIRY_VALUE"].quantile(0.25)
                expiry_value_avg = dte_filtered["EXPIRY_VALUE"].mean()
                expiry_value_q3 = dte_filtered["EXPIRY_VALUE"].quantile(0.75)
                
                # Append summary row
                summary_data = [
                    ["Number of Trades", num_trades],
                    ["Total Return", total_return],
                    ["Max Drawdown", max_drawdown],
                    ["Winrate (%)", winrate],
                    ["Risk/Reward Ratio", risk_reward_ratio],
                    ["Q1 Spread Cost", spread_cost_q1],
                    ["Avg Spread Cost", spread_cost_avg],
                    ["Q3 Spread Cost", spread_cost_q3],
                    ["Q1 Expiry Value", expiry_value_q1],
                    ["Avg Expiry Value", expiry_value_avg],
                    ["Q3 Expiry Value", expiry_value_q3]
                ]
                
                for row in summary_data:
                    ws.append(row)
                
                print(f"Updated {file}: DTE {dte_value} with strategy data.")
            except Exception as e:
                print(f"Error processing {file} - {sheet_name}: {e}")
        
        # Save workbook
        wb.save(file_path)
        wb.close()

if __name__ == "__main__":
    process_selected_files()
